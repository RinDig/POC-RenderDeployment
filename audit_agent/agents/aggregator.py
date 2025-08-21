"""
Aggregator Agent - Aggregates results and generates final report
"""

from typing import List
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

from ..core.base_agent import BaseAgent
from ..models.compliance_models import ComparisonResult, FinalReport
from ..utils.penalties import format_penalty_amount, get_audit_scope_disclaimer, get_excluded_penalties_context, DRC_MINING_PENALTIES


class AggregatorAgent(BaseAgent):
    """Aggregates results and generates final report"""
    
    def __init__(self, api_key: str = None):
        super().__init__("Aggregator", api_key=api_key)
    
    async def process(self, all_results: List[ComparisonResult]) -> FinalReport:
        """Aggregate all comparison results into final report"""
        print(f"[{self.name}] Aggregating {len(all_results)} comparison results")
        
        # Group by framework
        frameworks = list(set(result.framework for result in all_results))
        
        # Calculate overall compliance score
        total_score = sum(result.overall_score for result in all_results)
        overall_compliance_score = total_score / len(all_results) if all_results else 0.0
        
        # Calculate total financial exposure and penalties by framework
        total_max_penalty = 0.0
        penalty_by_framework = {}
        
        for result in all_results:
            if result.total_max_penalty_usd > 0:
                framework_key = result.framework
                if framework_key not in penalty_by_framework:
                    penalty_by_framework[framework_key] = 0.0
                penalty_by_framework[framework_key] += result.total_max_penalty_usd
                total_max_penalty += result.total_max_penalty_usd
        
        # Extract critical recommendations (score < 0.5) and add penalty info
        critical_recommendations = []
        for result in all_results:
            for item in result.items:
                if item.match_score < 0.5:
                    rec = f"[{result.framework}] {result.category}: {item.recommendation}"
                    # Add penalty info if applicable
                    if item.max_penalty_usd > 0:
                        rec += f" (Max Penalty: {format_penalty_amount(item.max_penalty_usd)})"
                    critical_recommendations.append(rec)
        
        # Generate executive summary including financial exposure with disclaimer
        summary_prompt = f"""
        Generate a concise executive summary (3-4 sentences) for this compliance audit:
        - Overall compliance score: {overall_compliance_score:.1%}
        - Frameworks assessed: {', '.join(frameworks)}
        - Critical gaps found: {len(critical_recommendations)}
        - Categories reviewed: {len(set(r.category for r in all_results))}
        - Total maximum financial exposure: {format_penalty_amount(total_max_penalty)}
        
        Note: Financial exposure includes administrative penalties only.
        Emphasize the financial risk if penalties are significant.
        """
        
        executive_summary = self.call_llm(
            summary_prompt,
            "You are an executive report writer for compliance audits."
        )
        
        # Add disclaimer to executive summary if DRC framework is included
        if any("DRC" in framework for framework in frameworks):
            disclaimer = get_audit_scope_disclaimer()
            executive_summary = f"{executive_summary}\n\n{disclaimer}"
        
        return FinalReport(
            timestamp=datetime.now().isoformat(),
            frameworks=frameworks,
            overall_compliance_score=overall_compliance_score,
            results=all_results,
            executive_summary=executive_summary,
            critical_recommendations=critical_recommendations[:10],  # Top 10
            total_max_penalty_usd=total_max_penalty,
            penalty_summary=penalty_by_framework
        )
    
    def generate_excel_report(self, report: FinalReport, output_path: str):
        """Generate Excel report from final report with professional formatting"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Executive Summary Sheet
            summary_data = {
                'Metric': ['Audit Date', 'Overall Compliance Score', 'Frameworks Assessed', 
                        'Total Categories', 'Critical Gaps', 'Total Maximum Financial Exposure', 
                        'Executive Summary'],
                'Value': [
                    report.timestamp,
                    f"{report.overall_compliance_score:.1%}",
                    ', '.join(report.frameworks),
                    len(set(r.category for r in report.results)),
                    len(report.critical_recommendations),
                    format_penalty_amount(report.total_max_penalty_usd),
                    report.executive_summary
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
            
            # Framework Summary Sheet
            framework_data = []
            for framework in report.frameworks:
                framework_results = [r for r in report.results if r.framework == framework]
                avg_score = sum(r.overall_score for r in framework_results) / len(framework_results) if framework_results else 0
                framework_penalty = report.penalty_summary.get(framework, 0.0)
                framework_data.append({
                    'Framework': framework,
                    'Categories Assessed': len(framework_results),
                    'Average Compliance': avg_score,
                    'Critical Gaps': sum(1 for r in framework_results for i in r.items if i.match_score < 0.5),
                    'Max Financial Exposure': format_penalty_amount(framework_penalty)
                })
            framework_df = pd.DataFrame(framework_data)
            framework_df.to_excel(writer, sheet_name='Framework Summary', index=False)
            
            # Detailed Findings Sheet
            detailed_data = []
            for result in report.results:
                for item in result.items:
                    row_data = {
                        'Framework': result.framework,
                        'Category': result.category,
                        'Requirement': item.question,
                        'Observation': item.input_statement,
                        'Reference': item.framework_ref,
                        'Compliance Score': item.match_score,
                        'Gap': item.gap,
                        'Recommendation': item.recommendation,
                        'Priority': 'Critical' if item.match_score < 0.5 else 'Medium' if item.match_score < 0.8 else 'Low'
                    }
                    
                    # Add penalty info if applicable
                    if item.potential_violations:
                        row_data['Violations'] = ', '.join([f"Art. {v}" for v in item.potential_violations])
                        row_data['Max Penalty'] = format_penalty_amount(item.max_penalty_usd)
                    else:
                        row_data['Violations'] = ''
                        row_data['Max Penalty'] = ''
                    
                    detailed_data.append(row_data)
            detailed_df = pd.DataFrame(detailed_data)
            detailed_df.to_excel(writer, sheet_name='Detailed Findings', index=False)
            
            # Critical Actions Sheet
            if report.critical_recommendations:
                actions_df = pd.DataFrame({
                    'Priority Action': report.critical_recommendations
                })
                actions_df.to_excel(writer, sheet_name='Critical Actions', index=False)
            
            # Financial Penalties Sheet (if applicable)
            if report.total_max_penalty_usd > 0:
                penalty_data = []
                
                # Add disclaimer row
                penalty_data.append({
                    'Article': 'DISCLAIMER',
                    'Violation': get_audit_scope_disclaimer(),
                    'Occurrences': '',
                    'Categories Affected': '',
                    'Max Fine (USD)': '',
                    'Applies To': ''
                })
                
                # Add excluded penalties context
                penalty_data.append({
                    'Article': 'EXCLUDED',
                    'Violation': get_excluded_penalties_context(),
                    'Occurrences': '',
                    'Categories Affected': '',
                    'Max Fine (USD)': '',
                    'Applies To': ''
                })
                
                # Add violations by article
                article_summary = {}
                for result in report.results:
                    for item in result.items:
                        if item.potential_violations:
                            for article in item.potential_violations:
                                if article not in article_summary:
                                    article_summary[article] = {
                                        'count': 0,
                                        'categories': set(),
                                        'max_penalty': 0
                                    }
                                article_summary[article]['count'] += 1
                                article_summary[article]['categories'].add(result.category)
                                
                for article, summary in sorted(article_summary.items()):
                    penalty_info = DRC_MINING_PENALTIES.get(article)
                    if penalty_info:
                        penalty_data.append({
                            'Article': f"Art. {article}",
                            'Violation': penalty_info.violation_description,
                            'Occurrences': summary['count'],
                            'Categories Affected': ', '.join(sorted(summary['categories'])),
                            'Max Fine (USD)': format_penalty_amount(penalty_info.max_fine_usd),
                            'Applies To': penalty_info.applies_to
                        })
                
                # Add total row
                penalty_data.append({
                    'Article': 'TOTAL',
                    'Violation': 'Maximum Financial Exposure',
                    'Occurrences': sum(s['count'] for s in article_summary.values()),
                    'Categories Affected': 'All',
                    'Max Fine (USD)': format_penalty_amount(report.total_max_penalty_usd),
                    'Applies To': 'Entity'
                })
                
                penalty_df = pd.DataFrame(penalty_data)
                penalty_df.to_excel(writer, sheet_name='Financial Penalties', index=False)
            
            # Get workbook and apply formatting
            workbook = writer.book
            
            # Color scheme
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            # Compliance score colors
            critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            warning_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            good_fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
            
            # Priority colors
            priority_colors = {
                'Critical': PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid"),
                'Medium': PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid"),
                'Low': PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
            }
            priority_fonts = {
                'Critical': Font(color="FFFFFF", bold=True),
                'Medium': Font(color="FFFFFF"),
                'Low': Font(color="FFFFFF")
            }
            
            # Border style
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Format Executive Summary Sheet
            summary_sheet = workbook['Executive Summary']
            
            # Headers
            for cell in summary_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Format compliance score cell
            for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row):
                if row[0].value == 'Overall Compliance Score':
                    score_value = float(row[1].value.strip('%')) / 100
                    if score_value < 0.5:
                        row[1].fill = critical_fill
                    elif score_value < 0.8:
                        row[1].fill = warning_fill
                    else:
                        row[1].fill = good_fill
                    row[1].font = Font(bold=True, size=14)
                
                # Apply borders
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Adjust column widths
            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 60
            
            # Format Framework Summary Sheet
            framework_sheet = workbook['Framework Summary']
            
            # Headers
            for cell in framework_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Format data rows
            for row in framework_sheet.iter_rows(min_row=2, max_row=framework_sheet.max_row):
                # Color code average compliance
                if isinstance(row[2].value, (int, float)):
                    score = row[2].value
                else:
                    score = float(row[2].value.strip('%')) / 100 if row[2].value else 0
                
                if score < 0.5:
                    row[2].fill = critical_fill
                    row[2].font = Font(color="FFFFFF", bold=True)
                elif score < 0.8:
                    row[2].fill = warning_fill
                    row[2].font = Font(bold=True)
                else:
                    row[2].fill = good_fill
                    row[2].font = Font(color="FFFFFF", bold=True)
                
                # Format percentage
                row[2].value = f"{score:.1%}"
                
                # Apply borders and alignment
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            framework_sheet.column_dimensions['A'].width = 20
            framework_sheet.column_dimensions['B'].width = 20
            framework_sheet.column_dimensions['C'].width = 20
            framework_sheet.column_dimensions['D'].width = 15
            
            # Format Detailed Findings Sheet
            detailed_sheet = workbook['Detailed Findings']
            
            # Headers
            for cell in detailed_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Format data rows
            for row in detailed_sheet.iter_rows(min_row=2, max_row=detailed_sheet.max_row):
                # Color code compliance score
                score = row[5].value if isinstance(row[5].value, (int, float)) else 0
                if score < 0.5:
                    row[5].fill = critical_fill
                    row[5].font = Font(color="FFFFFF", bold=True)
                elif score < 0.8:
                    row[5].fill = warning_fill
                    row[5].font = Font(bold=True)
                else:
                    row[5].fill = good_fill
                    row[5].font = Font(color="FFFFFF", bold=True)
                
                # Format percentage
                row[5].value = f"{score:.1%}"
                
                # Color code priority
                priority = row[8].value
                if priority in priority_colors:
                    row[8].fill = priority_colors[priority]
                    row[8].font = priority_fonts[priority]
                
                # Apply borders and alignment
                for i, cell in enumerate(row):
                    cell.border = thin_border
                    if i in [2, 3, 6, 7]:  # Long text columns
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            column_widths = {
                'A': 15,  # Framework
                'B': 20,  # Category
                'C': 40,  # Requirement
                'D': 40,  # Observation
                'E': 15,  # Reference
                'F': 15,  # Score
                'G': 35,  # Gap
                'H': 40,  # Recommendation
                'I': 10,  # Priority
                'J': 15,  # Violations
                'K': 20   # Max Penalty
            }
            for col, width in column_widths.items():
                detailed_sheet.column_dimensions[col].width = width
            
            # Format Critical Actions Sheet
            if 'Critical Actions' in workbook.sheetnames:
                actions_sheet = workbook['Critical Actions']
                
                # Header
                for cell in actions_sheet[1]:
                    cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True, size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # Format data rows
                for row in actions_sheet.iter_rows(min_row=2, max_row=actions_sheet.max_row):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        # Highlight with light red background
                        cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                
                # Adjust column width
                actions_sheet.column_dimensions['A'].width = 100
            
            # Format Financial Penalties Sheet
            if 'Financial Penalties' in workbook.sheetnames:
                penalties_sheet = workbook['Financial Penalties']
                
                # Header
                for cell in penalties_sheet[1]:
                    cell.fill = PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True, size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # Format data rows
                for row_idx, row in enumerate(penalties_sheet.iter_rows(min_row=2, max_row=penalties_sheet.max_row), start=2):
                    # Check if this is the TOTAL row
                    is_total_row = row[0].value == 'TOTAL'
                    
                    for cell in row:
                        cell.border = thin_border
                        
                        if is_total_row:
                            # Bold and highlight total row
                            cell.font = Font(bold=True, size=12)
                            cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                        
                        # Alignment based on column
                        if cell.column in [3, 5]:  # Numeric columns
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Adjust column widths
                penalties_sheet.column_dimensions['A'].width = 12   # Article
                penalties_sheet.column_dimensions['B'].width = 50   # Violation
                penalties_sheet.column_dimensions['C'].width = 15   # Occurrences
                penalties_sheet.column_dimensions['D'].width = 30   # Categories
                penalties_sheet.column_dimensions['E'].width = 20   # Max Fine
                penalties_sheet.column_dimensions['F'].width = 20   # Applies To
            
            print(f"[{self.name}] Excel report generated: {output_path}")